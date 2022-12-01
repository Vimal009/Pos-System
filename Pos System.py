from tkinter import *
from tkinter import messagebox
import openpyxl as xl
import datetime


window = Tk()
window.geometry('800x800')
window.title("Welcome")


stock_count_itm1 = IntVar()
stock_count_itm2 = IntVar()
stock_count_itm3 = IntVar()
stock_count_itm4 = IntVar()
stock_count_itm5 = IntVar()
stock_count_itm6 = IntVar()


var_1 = IntVar()
var_2 = IntVar()
var_3 = IntVar()
var_4 = IntVar()
var_5 = IntVar()
var_6 = IntVar()


price_item1 = 50
price_item2 = 35
price_item3 = 30
price_item4 = 20
price_item5 = 40
price_item6 = 15
total_cash = IntVar()
received_cash = IntVar()
balance_cash = StringVar()


def total():
    total_form = price_item1 * int(var_1.get()) + price_item2 * int(var_2.get()) + price_item3 * int(var_3.get()) + price_item4 * int(var_4.get()) + price_item5 * int(var_5.get()) + price_item6 * int(var_6.get())
    total_cash.set(total_form)


def exitt():
    exit()


def thank_you():
    messagebox.showinfo("Thank", "Thank you for shopping, Have a nice day :)")
    exit()


def calculate():
    temp = received_cash.get() - int(total_cash.get())
    balance_cash.set(temp)


def about():
    messagebox.showinfo('About', """
    Hi, 
    Name: M.J.Dhevaakar,
    Partnership = Vimal Shankar
    Education: First year Computer science in Pillai college, 
    Description: Owner and creator of this POS system.
    """)


# window 2 code starts here....


def win2():
    if int(var_1.get()) > quantity_1:
        messagebox.showerror('stock', f"Stock available is {quantity_1} for Onion. Please add stocks")

    elif int(var_2.get()) > quantity_2:
        messagebox.showerror('stock', f"Stock available is {quantity_2} for Potato. Please add stocks")

    elif int(var_3.get()) > quantity_3:
        messagebox.showerror('stock', f"Stock available is {quantity_3} for Tomato. Please add stocks")

    elif int(var_4.get()) > quantity_4:
        messagebox.showerror('stock', f"Stock available is {quantity_4} for Brinjal. Please add stocks")

    elif int(var_5.get()) > quantity_5:
        messagebox.showerror('stock', f"Stock available is {quantity_5} for Carrot. Please add stocks")

    elif int(var_6.get()) > quantity_6:
        messagebox.showerror('stock', f"Stock available is {quantity_6} for Cabbage. Please add stocks")

    else:
        window2 = Toplevel(window)
        window2.geometry('500x600')
        window.title("Cashout")
        butn2 = Button(window2, text='Cashout', bg='brown', fg='white', width=10, font=("TimesNewRoman", 12, 'bold'), command=total)
        butn2.place(x=350, y=50)
        Label(window2, text="Total: ", width=10, font=("TimesNewRoman", 18, 'bold')).place(x=20, y=50)
        Label(window2, text="", textvariable=total_cash, relief='raise', bg='white', width=10, font=("TimesNewRoman", 18, 'bold')).place(x=150, y=50)

        butn_end = Button(window2, text='End', bg='brown', fg='white', width=10, font=("TimesNewRoman", 12, 'bold'), command=thank_you)
        butn_end.place(x=250, y=350)
        total_cash.set(0)

        receive_cash = Label(window2, text="Pay:", width=10, font=("TimesNewRoman", 18, 'bold')).place(x=20, y=100)
        Entry(window2, textvar=received_cash, width=10, font=("TimesNewRoman", 18, 'bold')).place(x=150, y=100)
        butn_pay = Button(window2, text='Enter', width=6, font=("TimesNewRoman", 12, 'bold'), command=calculate).place(x=350, y=100)
        Label(window2, text="Balance: ", width=10, font=("TimesNewRoman", 18, 'bold')).place(x=20, y=150)
        bal = Label(window2, text="", textvariable=balance_cash, font=("TimesNewRoman", 18, 'bold')).place(x=200, y=150)
        received_cash.set(0)

        stock_count_itm1.set(quantity_1 - int(var_1.get()))
        stock_count_itm2.set(quantity_2 - int(var_2.get()))
        stock_count_itm3.set(quantity_3 - int(var_3.get()))
        stock_count_itm4.set(quantity_4 - int(var_4.get()))
        stock_count_itm5.set(quantity_5 - int(var_5.get()))
        stock_count_itm6.set(quantity_6 - int(var_6.get()))

        # stock

        wb = xl.load_workbook(file)
        sheet = wb.active
        sheet['b4'] = stock_count_itm1.get()
        sheet['b5'] = stock_count_itm2.get()
        sheet['b6'] = stock_count_itm3.get()
        sheet['b7'] = stock_count_itm4.get()
        sheet['b8'] = stock_count_itm5.get()
        sheet['b9'] = stock_count_itm6.get()
        wb.save(file)

        file1 = 'POS bill.xlsx'
        bi = xl.load_workbook(file1)
        sheet1 = bi.active


        # Bill

        file1 = 'POS bill.xlsx'
        bi = xl.load_workbook(file1)
        sheet1 = bi.active

        itm1 = 'Onion'
        itm2 = 'Potato'
        itm3 = 'Tomato'
        itm4 = 'Brinjal'
        itm5 = 'Carrot'
        itm6 = 'Cabbage'

        # bill description(items)

        sheet1['b11'].value = itm1
        sheet1['b12'].value = itm2
        sheet1['b13'].value = itm3
        sheet1['b14'].value = itm4
        sheet1['b15'].value = itm5
        sheet1['b16'].value = itm6

        sheet1['d11'] = var_1.get()
        sheet1['d12'] = var_2.get()
        sheet1['d13'] = var_3.get()
        sheet1['d14'] = var_4.get()
        sheet1['d15'] = var_5.get()
        sheet1['d16'] = var_6.get()

        # bill rate

        sheet1['e11'].value = price_item1
        sheet1['e12'].value = price_item2
        sheet1['e13'].value = price_item3
        sheet1['e14'].value = price_item4
        sheet1['e15'].value = price_item5
        sheet1['e16'].value = price_item6

        # bill amount
        sheet1['f11'] = var_1.get() * price_item1
        sheet1['f12'] = var_2.get() * price_item2
        sheet1['f13'] = var_3.get() * price_item3
        sheet1['f14'] = var_4.get() * price_item4
        sheet1['f15'] = var_5.get() * price_item5
        sheet1['f16'] = var_6.get() * price_item6

        # total amount in bill
        sheet1['f17'] = price_item1 * var_1.get() + price_item2 * var_2.get() + price_item3 * var_3.get() + price_item4 * var_4.get() + price_item5 * var_5.get() + price_item6 * var_6.get()

        # date and time

        sheet1['c8'] = datetime.date(2020, 2, 28)
        sheet1['c9'] = datetime.datetime.now()
        bi.save(file1)


# window 2 code ends here.....

label1 = Label(window, text="POINT OF SALE", width=50, height=2, bg='yellow', fg='blue', relief='solid', font=("TimesNewRoman", 25, 'bold')).pack()

sr_no = Label(window, text="Sr.no", bg='green', fg='yellow', width=8, relief='raise', font=("TimesNewRoman", 18, 'bold')).place(x=80, y=120)
lab_01 = Label(window, text='1.', bg='white', relief='sunken', width=5, font=("TimesNewRoman", 16, 'bold')).place(x=100, y=180)
lab_02 = Label(window, text='2.', bg='white', relief='sunken', width=5, font=("TimesNewRoman", 16, 'bold')).place(x=100, y=230)
lab_03 = Label(window, text='3.', bg='white', relief='sunken', width=5, font=("TimesNewRoman", 16, 'bold')).place(x=100, y=280)
lab_04 = Label(window, text='4.', bg='white', relief='sunken', width=5, font=("TimesNewRoman", 16, 'bold')).place(x=100, y=330)
lab_05 = Label(window, text='5.', bg='white', relief='sunken', width=5, font=("TimesNewRoman", 16, 'bold')).place(x=100, y=380)
lab_06 = Label(window, text='6.', bg='white', relief='sunken', width=5, font=("TimesNewRoman", 16, 'bold')).place(x=100, y=430)


price = Label(window, text="Price", bg='green', fg='yellow', relief='raise', width=8, font=("TimesNewRoman", 18, 'bold')).place(x=420, y=120)
price1 = Label(window, text="₹ 50.00", bg='white', relief='sunken', width=8, font=('TimesNewRoman', 16, 'bold')).place(x=420, y=180)
price2 = Label(window, text="₹ 35.00", bg='white', relief='sunken', width=8, font=('TimesNewRoman', 16, 'bold')).place(x=420, y=230)
price3 = Label(window, text="₹ 30.00", bg='white', relief='sunken', width=8, font=('TimesNewRoman', 16, 'bold')).place(x=420, y=280)
price4 = Label(window, text="₹ 20.00", bg='white', relief='sunken', width=8, font=('TimesNewRoman', 16, 'bold')).place(x=420, y=330)
price5 = Label(window, text="₹ 40.00", bg='white', relief='sunken', width=8, font=('TimesNewRoman', 16, 'bold')).place(x=420, y=380)
price6 = Label(window, text="₹ 15.00", bg='white', relief='sunken', width=8, font=('TimesNewRoman', 16, 'bold')).place(x=420, y=430)


item_name = Label(window, text='Items', bg='green', fg='yellow', relief='raise', width=8, font=("TimesNewRoman", 18, 'bold')).place(x=250, y=120)
item1 = Label(window, text='Onion', bg='white', relief='sunken', width=8, font=('TimesNewRoman', 16, 'bold')).place(x=250, y=180)
item2 = Label(window, text='Potato', bg='white', relief='sunken', width=8, font=('TimesNewRoman', 16, 'bold')).place(x=250, y=230)
item3 = Label(window, text='Tomato', bg='white', relief='sunken', width=8, font=('TimesNewRoman', 16, 'bold')).place(x=250, y=280)
item4 = Label(window, text='Brinjal', bg='white', relief='sunken', width=8, font=('TimesNewRoman', 16, 'bold')).place(x=250, y=330)
item5 = Label(window, text='Carrot', bg='white', relief='sunken', width=8, font=('TimesNewRoman', 16, 'bold')).place(x=250, y=380)
item6 = Label(window, text='Cabbage', bg='white', relief='sunken', width=8, font=('TimesNewRoman', 16, 'bold')).place(x=250, y=430)


quantity_name = Label(window, text='Quantity', bg='green', fg='yellow', width=8, relief='raise', font=("TimesNewRoman", 18, 'bold')).place(x=590, y=120)


list1 = [0, 1, 2, 5]
d1 = OptionMenu(window, var_1, *list1)
var_1.set("Select")
d1.config(width=8, bg='black', fg='white', font=("TimesNewRoman", 12, 'bold'))
d1.place(x=590, y=175)

d2 = OptionMenu(window, var_2, *list1)
var_2.set("Select")
d2.config(width=8, bg='black', fg='white', font=("TimesNewRoman", 12, 'bold'))
d2.place(x=590, y=225)

d3 = OptionMenu(window, var_3, *list1)
var_3.set("Select")
d3.config(width=8, bg='black', fg='white', font=("TimesNewRoman", 12, 'bold'))
d3.place(x=590, y=275)

d4 = OptionMenu(window, var_4, *list1)
var_4.set("Select")
d4.config(width=8, bg='black', fg='white', font=("TimesNewRoman", 12, 'bold'))
d4.place(x=590, y=325)

d5 = OptionMenu(window, var_5, *list1)
var_5.set("Select")
d5.config(width=8, bg='black', fg='white', font=("TimesNewRoman", 12, 'bold'))
d5.place(x=590, y=375)

d6 = OptionMenu(window, var_6, *list1)
var_6.set("Select")
d6.config(width=8, bg='black', fg='white', font=("TimesNewRoman", 12, 'bold'))
d6.place(x=590, y=425)






butn1 = Button(window, text='Exit', bg='brown', fg='white', width=10, font=("TimesNewRoman", 12, 'bold'), command=exitt)
butn1.place(x=250, y=650)

butn_next = Button(window, text='Next', bg='brown', fg='white', width=10, font=("TimesNewRoman", 12, 'bold'), command=win2)
butn_next.place(x=450, y=650)

menu = Menu(window)
window.config(menu=menu)

subm1 = Menu(menu)
menu.add_cascade(label='File', menu=subm1)
subm1.add_command(label='Exit', command=exit)

subm2 = Menu(menu)
menu.add_cascade(label='Option', menu=subm2)
subm2.add_command(label='About', command=about)





file = ("POS stock.xlsx")

wb = xl.load_workbook(file)
sheet = wb.active
quantity_1 = sheet['b4'].value
quantity_2 = sheet['b5'].value
quantity_3 = sheet['b6'].value
quantity_4 = sheet['b7'].value
quantity_5 = sheet['b8'].value
quantity_6 = sheet['b9'].value

wb.save(file)


mainloop()